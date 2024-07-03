from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def schematic(imgpath : str):
    base = Image.open(imgpath)
    return list(base.convert("RGB").getdata())

def rgbtohex(rgb : tuple):
    return "{:02x}{:02x}{:02x}".format(rgb[0], rgb[1], rgb[2]).upper()

def xlsximport(xlsxpath : str, scenario : list, imgpath : str):
    max_width, max_height = Image.open(imgpath).size
    pixel_counter = 0
    xcel = load_workbook(xlsxpath)
    xshe = xcel.active
    xshe.delete_rows(1, xshe.max_row+1)
    for row in range(1, max_height+1):
        for column in range(1, max_width+1):
            xshe[xshe.cell(row, column).coordinate].fill = PatternFill(start_color=rgbtohex(scenario[pixel_counter]), fill_type='solid')
            pixel_counter += 1
    xcel.save(xlsxpath)
 